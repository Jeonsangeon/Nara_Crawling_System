import openpyxl

def init_excel(option, notice = True):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = '설정옵션'
    if notice:
        worksheet.append(['업무', '공고/개찰일', '기관명'])
        for index, value in enumerate(option[0], 2):
            if not value:
                worksheet[f'A{index}'] = "전체"
            else:
                if value == 1:
                    worksheet[f'A{index}'] = "물품"
                if value == 3:
                    worksheet[f'A{index}'] = "공사"
                if value == 5:
                    worksheet[f'A{index}'] = "용역"
                if value == 6:
                    worksheet[f'A{index}'] = "리스"
                if value == 2:
                    worksheet[f'A{index}'] = "외자"
                if value == 11:
                    worksheet[f'A{index}'] = "비축"
                if value == 4:
                    worksheet[f'A{index}'] = "기타"
                if value == 20:
                    worksheet[f'A{index}'] = "민간"
        if option[1] == 1:
            worksheet['B2'] = '최근1개월'
        elif option[1] == 2:
            worksheet['B2'] = '최근3개월'
        else:
            worksheet['B2'] = '최근6개월'
        worksheet['C2'] = option[2]
        workbook.save(filename='입찰공고목록.xlsx')
    else:
        worksheet.append(['업무', '기관별검색', '품명(사업명)'])
        for index, value in enumerate(option[1], 2):
            if value == "A":
                worksheet[f'A{index}'] = "전체"
            elif value == 1:
                worksheet[f'A{index}'] = "물품"
            elif value == 3:
                worksheet[f'A{index}'] = "공사"
            elif value == 5:
                worksheet[f'A{index}'] = "용역"
            elif value == 2:
                worksheet[f'A{index}'] = "외자"
        if option[0][0] == 1:
            worksheet['B2'] = "공고기관: " + option[0][1]
        else:
            worksheet['B2'] = "수요기관: " + option[0][1]
        worksheet['C2'] = option[2]
        workbook.save(filename='사전규격목록.xlsx')
            
def make_announcement_sheet(announce_list, sheet_name):
    workbook = openpyxl.load_workbook('입찰공고목록.xlsx')
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(['업무', '분류', '공고명', '공고기관', '수요기관','공동수급', '입찰개시', '입찰마감', '개찰(입찰)', '입찰참가자격등록', '공동수급협정서마감', '사업금액(추정가격 + 부가세)', '추정금액', '추정가격', '부가가치세', '배정예산', '참가가능지역', "URL"])

    for row in announce_list:
        worksheet.append(row)

    workbook.save(filename='입찰공고목록.xlsx')

def make_prestandard_sheet(input_data):
    workbook = openpyxl.load_workbook('사전규격목록.xlsx')
    worksheet = workbook.create_sheet('목록')
    worksheet.append(['품명(사업명)', '배정예산액', '공개일시(나라장터 수신일시)', '의견등록마감일시', '공고기관', '수요기관', 'URL'])

    for data in input_data:
        worksheet.append(data)

    workbook.save(filename='사전규격목록.xlsx')